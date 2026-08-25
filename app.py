"""
Move the Cash — Scorecard Billing México
Streamlit app · replica de TABLERO_MEX_OPS_SCORE
"""

import streamlit as st
import openpyxl
import pandas as pd
from datetime import date, datetime
from collections import defaultdict
import io

# ── Config ────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="💰 Move the Cash",
    page_icon="💰",
    layout="wide",
)

BANXICO_FIX = 17.1092

# ── Helpers ──────────────────────────────────────────────────────────────────
def to_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date):     return v
    return None

def norm_team(t):
    if not t or str(t).strip() in ("", "—", "None"): return "—"
    words = str(t).strip().split()
    return " ".join(w if (w.isupper() and len(w) >= 2) else w.capitalize() for w in words)

def find_header(rows):
    return next(i for i, r in enumerate(rows)
                if any(str(v or "").strip() == "Job Number" for v in r))

def build_hmap(header_row):
    return {str(h).strip(): j for j, h in enumerate(header_row) if h}

def get(row, hm, col):
    i = hm.get(col)
    return row[i] if i is not None and i < len(row) else None

def semaforo(score):
    if score >= 0.98: return "🟢", "#166534", "#dcfce7"
    if score >= 0.90: return "🟡", "#854d0e", "#fef9c3"
    return "🔴", "#991b1b", "#fee2e2"

# ── Parsers ──────────────────────────────────────────────────────────────────
def parse_sheet(ws, use_wip_usd=False):
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = find_header(rows)
    hm = build_hmap(rows[hdr_idx])
    has_wip_usd = "WIP USD" in hm

    result, op_team_map = [], {}
    for row in rows[hdr_idx + 1:]:
        if not any(v is not None for v in row): continue

        bs = str(get(row, hm, "Billing Status (Criterion)") or
                 get(row, hm, "Overdue Status (Billing Criterion)") or "").strip()
        tier = "on_time" if bs == "On Time" else ("out_of_time" if bs == "Out of Time" else "no_data")

        etd = to_date(get(row, hm, "Origin ETD"))
        eta = to_date(get(row, hm, "Destination ETA"))
        op  = str(get(row, hm, "Operator (Preferred Full Name)") or "—").strip()

        if has_wip_usd and use_wip_usd:
            wip_usd = float(get(row, hm, "WIP USD") or 0)
        else:
            wip_usd = float(get(row, hm, "WIP (Local Currency)") or 0) / BANXICO_FIX

        raw_team = get(row, hm, "Equipo operativo") or get(row, hm, "Equipo Operativo") or ""
        team = norm_team(raw_team) if raw_team else "—"

        if team != "—" and op != "—":
            op_team_map[op] = team

        result.append({
            "job":      str(get(row, hm, "Job Number") or ""),
            "bs":       bs or "No Data",
            "tier":     tier,
            "wip_usd":  wip_usd,
            "operator": op,
            "team":     team,
            "has_date": bool(etd or eta),
        })
    return result, op_team_map

def load_pbr(uploaded_file):
    data = uploaded_file.read()
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    if "pending invoicing" in sheet_names_lower and "invoiced" in sheet_names_lower:
        ws_pnd = wb["Pending Invoicing"]
        ws_inv = wb["Invoiced"]
        pnd_rows, op_team_map = parse_sheet(ws_pnd, use_wip_usd=True)
        inv_rows, inv_team_map = parse_sheet(ws_inv, use_wip_usd=False)
        # Fill team for invoiced from pending map
        for r in inv_rows:
            if r["team"] == "—":
                r["team"] = op_team_map.get(r["operator"], "—")
    else:
        ws_pnd = wb.worksheets[0]
        ws_inv = wb.worksheets[1]
        pnd_rows, op_team_map = parse_sheet(ws_pnd, use_wip_usd=False)
        inv_rows, _ = parse_sheet(ws_inv, use_wip_usd=False)
        for r in inv_rows:
            if r["team"] == "—":
                r["team"] = op_team_map.get(r["operator"], "—")

    return pnd_rows, inv_rows

# ── Scoring ──────────────────────────────────────────────────────────────────
def compute_scores(inv_rows, pnd_rows):
    ops = defaultdict(lambda: {"inv": [], "pnd": []})
    for r in inv_rows: ops[r["operator"]]["inv"].append(r)
    for r in pnd_rows: ops[r["operator"]]["pnd"].append(r)

    results = []
    for op, d in ops.items():
        inv, pnd = d["inv"], d["pnd"]
        pnd_total = len(pnd)
        inv_total = len(inv)

        team = next((r["team"] for r in pnd if r["team"] != "—"),
                    next((r["team"] for r in inv if r["team"] != "—"), "—"))

        ot       = sum(1 for r in inv if r["tier"] == "on_time")
        otb      = ot / inv_total if inv_total else 0
        p_otb    = otb * 0.45

        pnd_oot    = [r for r in pnd if r["tier"] == "out_of_time"]
        n_vencidas = len(pnd_oot)
        p_fv       = 0.25 * (1 - n_vencidas / pnd_total) if pnd_total else 0.25

        pnd_con_wip  = sum(1 for r in pnd if r["wip_usd"] > 0)
        pnd_sin_wip  = sum(1 for r in pnd if r["wip_usd"] <= 0)
        wip_capture  = pnd_con_wip / pnd_total if pnd_total else 1.0
        p_wip        = wip_capture * 0.25

        pnd_at_risk     = [r for r in pnd if r["tier"] in ("out_of_time", "no_data")]
        wip_at_risk_usd = sum(r["wip_usd"] for r in pnd_at_risk)

        jobs_with_date = sum(1 for r in pnd if r["has_date"])
        dq    = jobs_with_date / pnd_total if pnd_total else 0
        p_dq  = dq * 0.05
        score = p_otb + p_fv + p_wip + p_dq

        results.append({
            "Operativo":       op,
            "Equipo":          team,
            "Jobs Facturados": inv_total,
            "OTB %":           otb,
            "Pond. OTB":       p_otb,
            "Jobs Pend. Venc": n_vencidas,
            "Pond. Venc":      p_fv,
            "Jobs sin WIP":    pnd_sin_wip,
            "WIP Capture %":   wip_capture,
            "Pond. WIP":       p_wip,
            "Jobs sin Fechas": pnd_total - jobs_with_date,
            "Pend. Total":     pnd_total,
            "Pond. DQ":        p_dq,
            "WIP Vencido USD": wip_at_risk_usd,
            "Score":           score,
            # raw for KPI
            "_ot": ot, "_pnd_total": pnd_total, "_con_wip": pnd_con_wip,
            "_with_date": jobs_with_date, "_inv_total": inv_total,
        })

    return sorted(results, key=lambda x: -x["Score"])

def compute_teams(results):
    teams = defaultdict(lambda: defaultdict(float))
    team_ops = defaultdict(list)
    for r in results:
        t = r["Equipo"]
        team_ops[t].append(r["Operativo"])
        for k in ["Jobs Facturados", "Jobs Pend. Venc", "Jobs sin WIP",
                   "Jobs sin Fechas", "WIP Vencido USD",
                   "_ot", "_pnd_total", "_con_wip", "_with_date", "_inv_total"]:
            teams[t][k] += r[k]

    team_results = []
    for t, d in teams.items():
        inv_total  = d["_inv_total"]
        pnd_total  = d["_pnd_total"]
        ot         = d["_ot"]
        con_wip    = d["_con_wip"]
        with_date  = d["_with_date"]
        n_venc     = d["Jobs Pend. Venc"]

        otb         = ot / inv_total if inv_total else 0
        wip_capture = con_wip / pnd_total if pnd_total else 1.0
        dq          = with_date / pnd_total if pnd_total else 0
        p_otb  = otb * 0.45
        p_fv   = 0.25 * (1 - n_venc / pnd_total) if pnd_total else 0.25
        p_wip  = wip_capture * 0.25
        p_dq   = dq * 0.05
        score  = p_otb + p_fv + p_wip + p_dq

        team_results.append({
            "Equipo":          t,
            "Operativos":      len(team_ops[t]),
            "Jobs Facturados": int(inv_total),
            "OTB %":           otb,
            "Pond. OTB":       p_otb,
            "Jobs Pend. Venc": int(n_venc),
            "Pond. Venc":      p_fv,
            "Jobs sin WIP":    int(d["Jobs sin WIP"]),
            "WIP Capture %":   wip_capture,
            "Pond. WIP":       p_wip,
            "Jobs sin Fechas": f"{int(d['Jobs sin Fechas'])} de {int(pnd_total)}",
            "Pond. DQ":        p_dq,
            "WIP Vencido USD": d["WIP Vencido USD"],
            "Score":           score,
        })
    return sorted(team_results, key=lambda x: -x["Score"])

# ── UI Helpers ───────────────────────────────────────────────────────────────
def score_color(score):
    if score >= 0.98: return "background-color:#dcfce7; color:#166534"
    if score >= 0.90: return "background-color:#fef9c3; color:#854d0e"
    return "background-color:#fee2e2; color:#991b1b"

def style_score_col(val):
    if val >= 0.98: return "background-color:#dcfce7; color:#166534; font-weight:700"
    if val >= 0.90: return "background-color:#fef9c3; color:#854d0e; font-weight:700"
    return "background-color:#fee2e2; color:#991b1b; font-weight:700"

def kpi_box(label, value, subtitle, color):
    st.markdown(f"""
    <div style="background:white;border-radius:12px;padding:16px 20px;
                border-left:4px solid {color};box-shadow:0 1px 4px rgba(0,0,0,.08)">
      <div style="font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase;
                  letter-spacing:1px;margin-bottom:4px">{label}</div>
      <div style="font-size:26px;font-weight:800;color:#0f172a">{value}</div>
      <div style="font-size:11px;color:#94a3b8;margin-top:4px">{subtitle}</div>
    </div>""", unsafe_allow_html=True)

# ── Main App ─────────────────────────────────────────────────────────────────
st.markdown("""
<div style="background:linear-gradient(135deg,#0f172a 0%,#1e3a5f 100%);
            border-radius:16px;padding:32px 40px;margin-bottom:24px">
  <div style="font-size:11px;font-weight:700;letter-spacing:3px;color:#60a5fa;
              text-transform:uppercase;margin-bottom:8px">xpd global · CargoWise</div>
  <h1 style="color:white;margin:0;font-size:32px;font-weight:900">💰 MOVE THE CASH</h1>
  <div style="color:#94a3b8;margin-top:6px;font-size:14px">
    Score de Cumplimiento por Talento · Billing México</div>
</div>""", unsafe_allow_html=True)

# ── File Upload ───────────────────────────────────────────────────────────────
uploaded = st.file_uploader(
    "Sube el archivo Mexico Billing Report (.xlsx)",
    type=["xlsx"],
    help="MEXICO_Billing_Report_*.xlsx o MEX_PBR_*.xlsx"
)

if not uploaded:
    st.info("👆 Sube el archivo Excel para generar el scorecard.")
    st.stop()

with st.spinner("Procesando datos..."):
    pnd_rows, inv_rows = load_pbr(uploaded)
    results     = compute_scores(inv_rows, pnd_rows)
    team_results = compute_teams(results)

df_op   = pd.DataFrame(results)
df_team = pd.DataFrame(team_results)

# ── Filters (sidebar) ─────────────────────────────────────────────────────────
st.sidebar.markdown("## 🔍 Filtros")
teams_list = sorted(df_op["Equipo"].unique())
sel_team = st.sidebar.selectbox("Equipo", ["Todos"] + [t for t in teams_list if t != "—"])
sel_sem  = st.sidebar.selectbox("Semáforo", ["Todos", "🟢 ≥98%", "🟡 90-97%", "🔴 <90%"])
search   = st.sidebar.text_input("Buscar operativo", "")

# Apply filters
df_filtered = df_op.copy()
if sel_team != "Todos":
    df_filtered = df_filtered[df_filtered["Equipo"] == sel_team]
if sel_sem == "🟢 ≥98%":
    df_filtered = df_filtered[df_filtered["Score"] >= 0.98]
elif sel_sem == "🟡 90-97%":
    df_filtered = df_filtered[(df_filtered["Score"] >= 0.90) & (df_filtered["Score"] < 0.98)]
elif sel_sem == "🔴 <90%":
    df_filtered = df_filtered[df_filtered["Score"] < 0.90]
if search:
    df_filtered = df_filtered[df_filtered["Operativo"].str.lower().str.contains(search.lower())]

# ── KPI Boxes ─────────────────────────────────────────────────────────────────
tot_inv   = df_filtered["_inv_total"].sum()
tot_ot    = df_filtered["_ot"].sum()
tot_pnd   = df_filtered["_pnd_total"].sum()
tot_venc  = df_filtered["Jobs Pend. Venc"].sum()
tot_con   = df_filtered["_con_wip"].sum()
tot_date  = df_filtered["_with_date"].sum()
tot_sinwip= df_filtered["Jobs sin WIP"].sum()
tot_wip_r = df_filtered["WIP Vencido USD"].sum()

otb_pct  = tot_ot / tot_inv * 100 if tot_inv else 0
fv_pct   = (1 - tot_venc / tot_pnd) * 100 if tot_pnd else 100
wip_pct  = tot_con / tot_pnd * 100 if tot_pnd else 100
dq_pct   = tot_date / tot_pnd * 100 if tot_pnd else 100

c1, c2, c3, c4, c5 = st.columns(5)
with c1:
    kpi_box("On Time Billing", f"{otb_pct:.1f}%",
            f"{int(tot_ot):,} de {int(tot_inv):,} facturados en tiempo", "#3b82f6")
with c2:
    kpi_box("Jobs Pend. Vencidos", f"{int(tot_venc):,}",
            f"{fv_pct:.1f}% cumplimiento · {int(tot_pnd):,} pend.", "#f59e0b")
with c3:
    kpi_box("WIP Capture", f"{wip_pct:.1f}%",
            f"{int(tot_con):,} de {int(tot_pnd):,} pend. con WIP", "#8b5cf6")
with c4:
    kpi_box("WIP Vencido", f"USD {tot_wip_r:,.0f}",
            "WIP acumulado · OOT + Sin Fecha", "#f97316")
with c5:
    kpi_box("Jobs sin Fechas", f"{int(tot_pnd - tot_date):,} de {int(tot_pnd):,}",
            f"{dq_pct:.1f}% con fechas registradas", "#10b981")

st.markdown("<br>", unsafe_allow_html=True)

# ── Tabs ──────────────────────────────────────────────────────────────────────
tab_team, tab_op = st.tabs(["📊 Por Equipo", "👤 Por Operativo"])

# ── Format helpers ────────────────────────────────────────────────────────────
def pct(v):   return f"{v*100:.1f}%"
def usd(v):   return f"USD {v:,.0f}"

# ── Tab Equipo ────────────────────────────────────────────────────────────────
with tab_team:
    disp_team = df_team.copy()
    disp_team["Sem"] = disp_team["Score"].apply(lambda s: semaforo(s)[0])
    disp_team["Score %"] = disp_team["Score"].apply(pct)
    disp_team["OTB %"]   = disp_team["OTB %"].apply(pct)
    disp_team["Pond. OTB"] = disp_team["Pond. OTB"].apply(pct)
    disp_team["WIP Capture %"] = disp_team["WIP Capture %"].apply(pct)
    disp_team["Pond. WIP"]   = disp_team["Pond. WIP"].apply(pct)
    disp_team["Pond. Venc"]  = disp_team["Pond. Venc"].apply(pct)
    disp_team["Pond. DQ"]    = disp_team["Pond. DQ"].apply(pct)
    disp_team["WIP Vencido USD"] = disp_team["WIP Vencido USD"].apply(usd)

    cols_team = ["Sem","Equipo","Operativos","Jobs Facturados",
                 "OTB %","Pond. OTB","Jobs Pend. Venc","Pond. Venc",
                 "Jobs sin WIP","WIP Capture %","Pond. WIP",
                 "Jobs sin Fechas","Pond. DQ","WIP Vencido USD","Score %"]

    st.dataframe(
        disp_team[cols_team],
        use_container_width=True,
        hide_index=True,
        height=400,
    )

# ── Tab Operativo ─────────────────────────────────────────────────────────────
with tab_op:
    disp_op = df_filtered.copy()
    disp_op["Sem"] = disp_op["Score"].apply(lambda s: semaforo(s)[0])
    disp_op["Score %"] = disp_op["Score"].apply(pct)
    disp_op["OTB %"]   = disp_op["OTB %"].apply(pct)
    disp_op["Pond. OTB"] = disp_op["Pond. OTB"].apply(pct)
    disp_op["WIP Capture %"] = disp_op["WIP Capture %"].apply(pct)
    disp_op["Pond. WIP"]  = disp_op["Pond. WIP"].apply(pct)
    disp_op["Pond. Venc"] = disp_op["Pond. Venc"].apply(pct)
    disp_op["Pond. DQ"]   = disp_op["Pond. DQ"].apply(pct)
    disp_op["WIP Vencido USD"] = disp_op["WIP Vencido USD"].apply(usd)
    disp_op["Jobs sin Fechas"] = disp_op.apply(
        lambda r: f"{r['Jobs sin Fechas']} de {r['Pend. Total']}", axis=1)

    cols_op = ["Sem","Operativo","Equipo","Jobs Facturados",
               "OTB %","Pond. OTB","Jobs Pend. Venc","Pond. Venc",
               "Jobs sin WIP","WIP Capture %","Pond. WIP",
               "Jobs sin Fechas","Pond. DQ","WIP Vencido USD","Score %"]

    st.dataframe(
        disp_op[cols_op],
        use_container_width=True,
        hide_index=True,
        height=500,
    )

    # Resumen
    n_green  = (df_filtered["Score"] >= 0.98).sum()
    n_yellow = ((df_filtered["Score"] >= 0.90) & (df_filtered["Score"] < 0.98)).sum()
    n_red    = (df_filtered["Score"] < 0.90).sum()
    st.caption(f"🟢 {n_green} · 🟡 {n_yellow} · 🔴 {n_red} · Total: {len(df_filtered)} operativos")

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown("---")
st.caption("💰 Move the Cash · xpd global · CargoWise Data Enabling")
